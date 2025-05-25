from django.shortcuts import render
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum, F, Count
from datetime import datetime, timedelta
from decimal import Decimal
from django.http import HttpResponse
import openpyxl
from reportlab.pdfgen import canvas
from django.utils.dateparse import parse_date

from .models import Categoria, UnidadMedida, Producto, Movimiento
from .serializers import (
    CategoriaSerializer, UnidadMedidaSerializer,
    ProductoSerializer, MovimientoSerializer
)

def index(request):
    return render(request, 'inventario/index.html')

class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.annotate(
        productos_count=Count('producto'),
        stock_total=Sum('producto__stock_actual')
    )
    serializer_class = CategoriaSerializer

class UnidadMedidaViewSet(viewsets.ModelViewSet):
    queryset = UnidadMedida.objects.all()
    serializer_class = UnidadMedidaSerializer

class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    
    @action(detail=False, methods=['get'])
    def stock_bajo(self, request):
        """Retorna productos con stock bajo (menor al mínimo)"""
        productos = self.queryset.filter(
            stock_actual__lt=F('stock_minimo'),
            estado='activo'
        )
        serializer = self.get_serializer(productos, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def valor_inventario(self, request):
        """Calcula el valor total del inventario"""
        total = self.queryset.aggregate(
            total=Sum(F('stock_actual') * F('precio_compra'))
        )['total'] or Decimal('0')
        return Response({'valor_total': total})

class MovimientoViewSet(viewsets.ModelViewSet):
    queryset = Movimiento.objects.all()
    serializer_class = MovimientoSerializer
    
    @action(detail=False, methods=['get'])
    def reporte_movimientos(self, request):
        """Genera reporte de movimientos por período"""
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        tipo = request.query_params.get('tipo')
        
        movimientos = self.queryset
        
        if fecha_inicio:
            movimientos = movimientos.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            movimientos = movimientos.filter(fecha__lte=fecha_fin)
        if tipo:
            movimientos = movimientos.filter(tipo=tipo)
            
        serializer = self.get_serializer(movimientos, many=True)
        return Response(serializer.data)
    
    def create(self, request, *args, **kwargs):
        """Crear un nuevo movimiento y actualizar stock"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        
        # Actualizar estadísticas si es necesario
        producto = serializer.instance.producto
        if producto.stock_actual < producto.stock_minimo:
            # Aquí podrías enviar una notificación
            pass
            
        return Response(serializer.data, status=status.HTTP_201_CREATED)

def reporte_inventario_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.pdf"'
    p = canvas.Canvas(response)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 800, "Reporte de Inventario")
    p.setFont("Helvetica", 12)
    y = 770
    p.drawString(100, y, "Nombre")
    p.drawString(300, y, "Stock")
    p.drawString(400, y, "Precio Compra")
    p.drawString(500, y, "Precio Venta")
    y -= 20
    for producto in Producto.objects.all():
        p.drawString(100, y, producto.nombre)
        p.drawString(300, y, str(producto.stock_actual))
        p.drawString(400, y, str(producto.precio_compra))
        p.drawString(500, y, str(producto.precio_venta))
        y -= 20
        if y < 50:
            p.showPage()
            y = 800
    p.showPage()
    p.save()
    return response

def reporte_inventario_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Nombre', 'Stock', 'Precio Compra', 'Precio Venta'])
    for producto in Producto.objects.all():
        ws.append([producto.nombre, producto.stock_actual, producto.precio_compra, producto.precio_venta])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.xlsx"'
    wb.save(response)
    return response

# --- Reportes PDF y Excel ---

def stock_bajo_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="stock_bajo.pdf"'
    p = canvas.Canvas(response)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 800, "Reporte: Stock Bajo")
    p.setFont("Helvetica", 12)
    y = 770
    p.drawString(100, y, "Nombre")
    p.drawString(300, y, "Stock Actual")
    p.drawString(400, y, "Stock Mínimo")
    y -= 20
    productos = Producto.objects.filter(stock_actual__lt=F('stock_minimo'), estado='activo')
    for producto in productos:
        p.drawString(100, y, producto.nombre)
        p.drawString(300, y, str(producto.stock_actual))
        p.drawString(400, y, str(producto.stock_minimo))
        y -= 20
        if y < 50:
            p.showPage()
            y = 800
    p.showPage()
    p.save()
    return response

def stock_bajo_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Nombre', 'Stock Actual', 'Stock Mínimo'])
    productos = Producto.objects.filter(stock_actual__lt=F('stock_minimo'), estado='activo')
    for producto in productos:
        ws.append([producto.nombre, producto.stock_actual, producto.stock_minimo])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="stock_bajo.xlsx"'
    wb.save(response)
    return response

def valor_inventario_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="valor_inventario.pdf"'
    p = canvas.Canvas(response)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 800, "Reporte: Valor de Inventario")
    p.setFont("Helvetica", 12)
    total = Producto.objects.aggregate(total=Sum(F('stock_actual') * F('precio_compra')))['total'] or Decimal('0')
    p.drawString(100, 770, f"Valor total del inventario: ${total:.2f}")
    p.showPage()
    p.save()
    return response

def valor_inventario_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Valor Total'])
    total = Producto.objects.aggregate(total=Sum(F('stock_actual') * F('precio_compra')))['total'] or Decimal('0')
    ws.append([float(total)])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="valor_inventario.xlsx"'
    wb.save(response)
    return response

def movimientos_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="movimientos.pdf"'
    p = canvas.Canvas(response)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 800, "Reporte: Movimientos de Inventario")
    p.setFont("Helvetica", 12)
    y = 770
    p.drawString(100, y, "Fecha")
    p.drawString(200, y, "Tipo")
    p.drawString(300, y, "Producto")
    p.drawString(450, y, "Cantidad")
    y -= 20
    movimientos = Movimiento.objects.all()
    for mov in movimientos:
        p.drawString(100, y, mov.fecha.strftime('%Y-%m-%d'))
        p.drawString(200, y, mov.tipo)
        p.drawString(300, y, mov.producto.nombre)
        p.drawString(450, y, str(mov.cantidad))
        y -= 20
        if y < 50:
            p.showPage()
            y = 800
    p.showPage()
    p.save()
    return response

def movimientos_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="movimientos.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Fecha', 'Tipo', 'Producto', 'Cantidad'])
    movimientos = Movimiento.objects.all()
    for mov in movimientos:
        ws.append([mov.fecha.strftime('%Y-%m-%d'), mov.tipo, mov.producto.nombre, mov.cantidad])
    wb.save(response)
    return response 